import io
import re
from datetime import date, datetime

from flask import jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REFERENCE_SHEETS = {'範例', '內海巡禮-同業紀錄', '追風音樂祭', '機票與夢想民宿價格', '景點參考'}
SUPPORTING_MARKERS = ('名單', '費用表', '時間安排', '價格', '紀錄', '參考')
STANDARD_SHEET = '接團匯入'
OPERATION_FIELDS = (
    'group_name', 'source_sheet', 'contact_channel', 'sales_owner',
    'outbound_transport', 'return_transport', 'accommodation_details',
    'rooming_details', 'special_requirements', 'payment_status',
    'deposit_amount', 'balance_amount', 'supplier_notes',
)
STANDARD_COLUMNS = (
    ('系統編號', 'trip_id', '既有行程的系統編號；留白會建立新行程', 12),
    ('團名', 'group_name', '必填；團體或主要聯絡人名稱', 24),
    ('聯絡電話', 'customer_phone', '客戶聯絡電話', 16),
    ('聯絡Email', 'customer_email', '客戶電子郵件', 26),
    ('出發日期', 'trip_date', '格式：YYYY-MM-DD', 14),
    ('回程日期', 'return_date', '格式：YYYY-MM-DD', 14),
    ('成人', 'adults', '成人數量', 10),
    ('兒童', 'children', '兒童數量', 10),
    ('敬老', 'seniors', '敬老人數', 10),
    ('狀態', 'status', '草稿、報價中、確認或取消', 12),
    ('交通成本', 'transport_cost', '整團交通成本', 14),
    ('住宿成本', 'accommodation_cost', '整團住宿成本', 14),
    ('行程成本', 'activity_cost', '景點與活動成本', 14),
    ('餐食成本', 'meal_cost', '整團餐食成本', 14),
    ('其他成本', 'other_cost', '其他成本', 14),
    ('服務費', 'service_fee', '整團服務費', 12),
    ('聯絡來源', 'contact_channel', '例如 LINE、電話、同業', 16),
    ('承辦人', 'sales_owner', '負責此團的人員', 14),
    ('去程交通／航班', 'outbound_transport', '去程班次、時間或車輛', 28),
    ('回程交通／航班', 'return_transport', '回程班次、時間或車輛', 28),
    ('住宿資訊', 'accommodation_details', '住宿名稱、日期與聯絡資訊', 28),
    ('房型分配', 'rooming_details', '房型與間數', 24),
    ('特殊需求', 'special_requirements', '飲食、無障礙、兒童等需求', 28),
    ('付款狀態', 'payment_status', '未確認、待核對、已付訂金或已結清', 16),
    ('訂金', 'deposit_amount', '已付訂金金額', 12),
    ('尾款', 'balance_amount', '待付尾款金額', 12),
    ('供應商／訂位備註', 'supplier_notes', '訂位編號、窗口與供應商備註', 32),
    ('系統備註', 'notes', '行程內部備註', 32),
)
HEADER_TO_KEY = {header: key for header, key, _, _ in STANDARD_COLUMNS}
NUMBER_KEYS = {
    'trip_id', 'adults', 'children', 'seniors', 'transport_cost',
    'accommodation_cost', 'activity_cost', 'meal_cost', 'other_cost',
    'service_fee', 'deposit_amount', 'balance_amount',
}


def _safe_int(value, default=0):
    try:
        return int(float(value)) if value not in (None, '') else default
    except (TypeError, ValueError):
        return default


def _iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or '').strip()
    if not text:
        return ''
    for pattern in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text[:20]


def _excel_safe(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return "'" + value
    return value


def _date_from_parts(year, compact):
    compact = str(compact)
    if len(compact) == 2:
        month, day = int(compact[0]), int(compact[1])
    elif len(compact) == 3:
        month, day = int(compact[0]), int(compact[1:])
    elif len(compact) == 4:
        month, day = int(compact[:2]), int(compact[2:])
    else:
        return None
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_sheet_dates(title, default_year=None):
    year = default_year or datetime.utcnow().year
    normalized = title.replace('/', '').replace('.', '').replace('月', '').replace('日', '')
    full = re.search(r'(?P<year>20\d{2})(?P<start>\d{4})\s*[-~到]\s*(?P<end>\d{3,4})', normalized)
    if full:
        year = int(full.group('year'))
        start = _date_from_parts(year, full.group('start'))
        end = _date_from_parts(year, full.group('end'))
    else:
        pair = re.search(r'(?<!\d)(?P<start>\d{3,4})\s*[-~到]\s*(?P<end>\d{2,4})(?!\d)', normalized)
        if not pair:
            return '', ''
        start_text, end_text = pair.group('start'), pair.group('end')
        start = _date_from_parts(year, start_text)
        if len(end_text) <= 2 and start:
            end = (date(year, start.month, int(end_text))
                   if int(end_text) <= 31 else _date_from_parts(year, end_text))
        else:
            end = _date_from_parts(year, end_text)
    if not start or not end:
        return '', ''
    if end < start:
        try:
            end = end.replace(year=end.year + 1)
        except ValueError:
            return '', ''
    return start.isoformat(), end.isoformat()


def clean_group_name(title):
    name = re.sub(r'20\d{2}\d{4}\s*[-~到]\s*\d{3,4}', '', title)
    name = re.sub(r'(?<!\d)\d{3,4}\s*[-~到]\s*\d{2,4}(?!\d)', '', name)
    name = re.sub(r'[-_ ]*取消$', '', name)
    name = re.sub(r'^[-_ ]+|[-_ ]+$', '', name)
    return name or title


def estimate_people(title, text):
    for source in (title, text[:3000]):
        match = re.search(r'[Xx×]\s*(\d{1,3})\s*(?:人|位)?|(?<!\d)(\d{1,3})\s*(?:人|位)', source)
        if match:
            return max(1, min(500, int(match.group(1) or match.group(2))))
    return 2


def sheet_signals(text):
    checks = {
        '含每日行程': ('Day', '行程', '景點/內容'),
        '含航班資訊': ('航班', '機票', '航空'),
        '含車輛資訊': ('租車', '遊覽車', '機車', '車輛'),
        '含住宿資訊': ('住宿', '民宿', '飯店', '酒店'),
        '含房型資訊': ('房型', '雙人房', '四人房', '包棟'),
        '含付款資訊': ('已付款', '未付款', '訂金', '尾款'),
        '含訂妥註記': ('訂好了', '已訂', '訂位'),
    }
    return [label for label, words in checks.items() if any(word in text for word in words)]


def classify_sheet(title, trip_date):
    if title in REFERENCE_SHEETS:
        return 'reference'
    if any(marker in title for marker in SUPPORTING_MARKERS):
        return 'supporting'
    if trip_date:
        return 'cancelled' if '取消' in title else 'trip'
    return 'review'


def _standard_preview(worksheet):
    headers = [str(cell.value or '').strip() for cell in worksheet[1]]
    indexes = {HEADER_TO_KEY[header]: index for index, header in enumerate(headers) if header in HEADER_TO_KEY}
    if 'group_name' not in indexes:
        raise ValueError('標準工作表缺少「團名」欄位')
    items = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {key: row[index] if index < len(row) else None for key, index in indexes.items()}
        if not any(value not in (None, '') for value in values.values()):
            continue
        item = {key: (_safe_int(value) if key in NUMBER_KEYS else str(value or '').strip())
                for key, value in values.items()}
        item['trip_date'] = _iso_date(values.get('trip_date'))
        item['return_date'] = _iso_date(values.get('return_date'))
        item['group_name'] = str(values.get('group_name') or '').strip()
        item['status'] = str(values.get('status') or '草稿').strip()
        item['adults'] = _safe_int(values.get('adults'), 2)
        item['children'] = _safe_int(values.get('children'))
        item['seniors'] = _safe_int(values.get('seniors'))
        item['days'] = 2
        if item['trip_date'] and item['return_date']:
            try:
                item['days'] = (date.fromisoformat(item['return_date']) - date.fromisoformat(item['trip_date'])).days + 1
            except ValueError:
                pass
        kind = 'cancelled' if item['status'] == '取消' else ('trip' if item['group_name'] else 'review')
        item.update({
            'selected': kind == 'trip',
            'kind': kind,
            'source_sheet': str(values.get('source_sheet') or f'{STANDARD_SHEET} 第{row_number}列')[:200],
            'signals': ['標準格式'],
            'warning': '' if item['group_name'] and item['trip_date'] else '請確認團名與出發日期',
        })
        items.append(item)
    return items


def preview_workbook(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    items = []
    try:
        if STANDARD_SHEET in workbook.sheetnames:
            return _standard_preview(workbook[STANDARD_SHEET])
        for worksheet in workbook.worksheets:
            cells = []
            for row in worksheet.iter_rows(max_row=min(worksheet.max_row or 1, 120),
                                           max_col=min(worksheet.max_column or 1, 30),
                                           values_only=True):
                cells.extend(str(value) for value in row if value not in (None, ''))
            text = '\n'.join(cells)
            trip_date, return_date = parse_sheet_dates(worksheet.title, 2026)
            kind = classify_sheet(worksheet.title, trip_date)
            people = estimate_people(worksheet.title, text)
            items.append({
                'selected': kind == 'trip', 'source_sheet': worksheet.title,
                'group_name': clean_group_name(worksheet.title), 'trip_date': trip_date,
                'return_date': return_date,
                'days': ((date.fromisoformat(return_date) - date.fromisoformat(trip_date)).days + 1)
                        if trip_date and return_date else 2,
                'adults': people, 'children': 0, 'seniors': 0,
                'status': '取消' if kind == 'cancelled' else '草稿',
                'kind': kind, 'signals': sheet_signals(text),
                'warning': '' if trip_date else '未能從工作表名稱辨識日期',
            })
    finally:
        workbook.close()
    return items


def build_standard_workbook(rows=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = STANDARD_SHEET
    headers = [column[0] for column in STANDARD_COLUMNS]
    sheet.append(headers)
    for row in rows or []:
        sheet.append([_excel_safe(row.get(key, '')) for _, key, _, _ in STANDARD_COLUMNS])
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:AB{max(2, sheet.max_row)}'
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill('solid', fgColor='0F766E')
    for index, (_, _, _, width) in enumerate(STANDARD_COLUMNS, start=1):
        cell = sheet.cell(1, index)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions[cell.column_letter].width = width
    sheet.row_dimensions[1].height = 28
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for column in ('E', 'F'):
        for cell in sheet[column][1:]:
            cell.number_format = 'yyyy-mm-dd'
    status_validation = DataValidation(type='list', formula1='"草稿,報價中,確認,取消"', allow_blank=True)
    payment_validation = DataValidation(type='list', formula1='"未確認,待核對,已付訂金,已結清"', allow_blank=True)
    sheet.add_data_validation(status_validation)
    sheet.add_data_validation(payment_validation)
    status_validation.add(f'J2:J1000')
    payment_validation.add(f'X2:X1000')
    yellow = PatternFill('solid', fgColor='FEF3C7')
    sheet.conditional_formatting.add('B2:B1000', FormulaRule(formula=['B2=""'], fill=yellow))
    sheet.conditional_formatting.add('E2:E1000', FormulaRule(formula=['E2=""'], fill=yellow))

    guide = workbook.create_sheet('欄位說明')
    guide.append(['欄位', '是否必填', '說明'])
    for header, key, description, _ in STANDARD_COLUMNS:
        guide.append([header, '必填' if key in ('group_name', 'trip_date') else '選填', description])
    guide.freeze_panes = 'A2'
    guide.sheet_view.showGridLines = False
    guide.column_dimensions['A'].width = 24
    guide.column_dimensions['B'].width = 12
    guide.column_dimensions['C'].width = 58
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
    for row in guide.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    return workbook


def _workbook_response(workbook, filename):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def register_group_import(app, db, Trip, require_role, log_activity):
    class TripOperation(db.Model):
        __tablename__ = 'trip_operations'

        id = db.Column(db.Integer, primary_key=True)
        trip_id = db.Column(db.Integer, db.ForeignKey('trips.id', ondelete='CASCADE'), nullable=False, unique=True, index=True)
        group_name = db.Column(db.String(200), default='')
        source_sheet = db.Column(db.String(200), default='')
        contact_channel = db.Column(db.String(100), default='')
        sales_owner = db.Column(db.String(100), default='')
        outbound_transport = db.Column(db.Text, default='')
        return_transport = db.Column(db.Text, default='')
        accommodation_details = db.Column(db.Text, default='')
        rooming_details = db.Column(db.Text, default='')
        special_requirements = db.Column(db.Text, default='')
        payment_status = db.Column(db.String(50), default='未確認')
        deposit_amount = db.Column(db.Integer, default=0)
        balance_amount = db.Column(db.Integer, default=0)
        supplier_notes = db.Column(db.Text, default='')
        updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

        def to_dict(self):
            return {field: getattr(self, field) or (0 if field.endswith('_amount') else '')
                    for field in OPERATION_FIELDS}

    def operation_for(trip_id):
        return TripOperation.query.filter_by(trip_id=trip_id).first()

    @app.post('/api/trips/import/preview')
    @require_role('admin')
    def preview_trip_import():
        upload = request.files.get('file')
        if not upload or not upload.filename.lower().endswith('.xlsx'):
            return jsonify({'error': '請上傳 .xlsx Excel 檔'}), 400
        try:
            items = preview_workbook(upload.read())
        except Exception:
            return jsonify({'error': 'Excel 檔案無法解析，請確認格式或重新下載標準範本'}), 400
        counts = {kind: sum(1 for item in items if item['kind'] == kind)
                  for kind in ('trip', 'cancelled', 'reference', 'supporting', 'review')}
        return jsonify({'items': items, 'counts': counts, 'total_sheets': len(items)})

    @app.get('/api/trips/export/template')
    @require_role('admin', 'viewer')
    def export_trip_template():
        return _workbook_response(build_standard_workbook(), '澎湖接團匯入範本.xlsx')

    @app.get('/api/trips/export')
    @require_role('admin', 'viewer')
    def export_trips():
        trips = Trip.query.order_by(Trip.trip_date, Trip.id).all()
        operations = {item.trip_id: item for item in TripOperation.query.all()}
        rows = []
        for trip in trips:
            operation = operations.get(trip.id)
            op = operation.to_dict() if operation else {}
            rows.append({
                'trip_id': trip.id, 'group_name': op.get('group_name') or trip.customer_name,
                'customer_phone': trip.customer_phone, 'customer_email': trip.customer_email,
                'trip_date': trip.trip_date, 'return_date': trip.return_date,
                'adults': trip.adults, 'children': trip.children, 'seniors': trip.seniors,
                'status': trip.status, 'transport_cost': trip.transport_cost,
                'accommodation_cost': trip.accommodation_cost, 'activity_cost': trip.activity_cost,
                'meal_cost': trip.meal_cost, 'other_cost': trip.other_cost,
                'service_fee': trip.service_fee, 'notes': trip.notes, **op,
            })
        filename = f'澎湖接團資料-{datetime.now().strftime("%Y%m%d")}.xlsx'
        return _workbook_response(build_standard_workbook(rows), filename)

    @app.post('/api/trips/import')
    @require_role('admin')
    def import_trips():
        payload = request.get_json(silent=True) or {}
        items = payload.get('items')
        if not isinstance(items, list):
            return jsonify({'error': 'items 必須是陣列'}), 400
        created, updated, skipped = [], [], []
        for item in items[:1000]:
            if not item.get('selected'):
                continue
            source_sheet = str(item.get('source_sheet', '')).strip()[:200]
            trip_id = _safe_int(item.get('trip_id'))
            trip = db.session.get(Trip, trip_id) if trip_id else None
            if trip_id and not trip:
                skipped.append({'source_sheet': source_sheet, 'reason': f'找不到系統編號 {trip_id}'})
                continue
            if not trip and source_sheet:
                existing = TripOperation.query.filter_by(source_sheet=source_sheet).first()
                if existing:
                    skipped.append({'source_sheet': source_sheet, 'reason': '已匯入'})
                    continue
            is_new = trip is None
            trip = trip or Trip(itinerary_data='{}')
            if is_new:
                db.session.add(trip)
            trip.customer_name = str(item.get('group_name', '')).strip()[:100]
            trip.customer_phone = str(item.get('customer_phone', '')).strip()[:20]
            trip.customer_email = str(item.get('customer_email', '')).strip()[:100]
            trip.trip_date = _iso_date(item.get('trip_date'))
            trip.return_date = _iso_date(item.get('return_date'))
            trip.days = max(1, min(60, _safe_int(item.get('days'), 2)))
            trip.adults = max(0, min(500, _safe_int(item.get('adults'), 2)))
            trip.children = max(0, min(500, _safe_int(item.get('children'))))
            trip.seniors = max(0, min(500, _safe_int(item.get('seniors'))))
            trip.total_people = trip.adults + trip.children + trip.seniors
            for key in ('transport_cost', 'accommodation_cost', 'activity_cost', 'meal_cost', 'other_cost', 'service_fee'):
                if key in item:
                    setattr(trip, key, max(0, _safe_int(item.get(key))))
            trip.cost_subtotal = sum((getattr(trip, key, 0) or 0) for key in
                                     ('transport_cost', 'accommodation_cost', 'activity_cost',
                                      'meal_cost', 'other_cost'))
            trip.final_quote = trip.cost_subtotal + (trip.service_fee or 0)
            trip.quote_per_person = int(trip.final_quote / trip.total_people) if trip.total_people else 0
            trip.status = str(item.get('status') or '草稿')[:20]
            if 'notes' in item:
                trip.notes = str(item.get('notes') or '')[:4000]
            elif is_new:
                trip.notes = f'由 Excel 工作表「{source_sheet}」匯入，請確認成本與行程內容。'
            trip.updated_at = datetime.utcnow()
            db.session.flush()

            operation = operation_for(trip.id) or TripOperation(trip_id=trip.id)
            db.session.add(operation)
            operation.group_name = trip.customer_name
            operation.source_sheet = source_sheet or operation.source_sheet
            for field in OPERATION_FIELDS:
                if field not in item or field in ('group_name', 'source_sheet'):
                    continue
                if field.endswith('_amount'):
                    setattr(operation, field, max(0, _safe_int(item.get(field))))
                else:
                    setattr(operation, field, str(item.get(field) or '').strip()[:4000])
            log_activity('update' if not is_new else 'create', 'trip', f'Excel 匯入：{trip.customer_name}')
            target = {'id': trip.id, 'source_sheet': source_sheet}
            (created if is_new else updated).append(target)
        db.session.commit()
        return jsonify({
            'created': created, 'updated': updated, 'skipped': skipped,
            'created_count': len(created), 'updated_count': len(updated),
        }), 201

    @app.get('/api/trips/<int:trip_id>/operations')
    @require_role('admin', 'viewer')
    def get_trip_operations(trip_id):
        trip = db.get_or_404(Trip, trip_id)
        operation = operation_for(trip.id)
        return jsonify(operation.to_dict() if operation else {
            **{field: 0 if field.endswith('_amount') else '' for field in OPERATION_FIELDS},
            'group_name': trip.customer_name,
        })

    @app.put('/api/trips/<int:trip_id>/operations')
    @require_role('admin')
    def update_trip_operations(trip_id):
        trip = db.get_or_404(Trip, trip_id)
        data = request.get_json(silent=True) or {}
        operation = operation_for(trip.id) or TripOperation(trip_id=trip.id)
        db.session.add(operation)
        for field in OPERATION_FIELDS:
            if field not in data:
                continue
            if field.endswith('_amount'):
                setattr(operation, field, max(0, _safe_int(data[field])))
            else:
                limit = 200 if field in ('group_name', 'source_sheet') else 4000
                setattr(operation, field, str(data[field]).strip()[:limit])
        if operation.group_name:
            trip.customer_name = operation.group_name[:100]
        trip.updated_at = datetime.utcnow()
        log_activity('update', 'trip', f'團務資料：{trip.customer_name}')
        db.session.commit()
        return jsonify(operation.to_dict())

    return TripOperation
