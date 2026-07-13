import io
import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

_temp_dir = tempfile.TemporaryDirectory()
os.environ['DATABASE_URL'] = f"sqlite:///{os.path.join(_temp_dir.name, 'group-import.db')}"
os.environ['SECRET_KEY'] = 'test-secret'

from backend.app import ActivityLog, Trip, TripOperation, User, app, db  # noqa: E402
from backend.group_import import STANDARD_SHEET, parse_sheet_dates, preview_workbook  # noqa: E402


def legacy_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '0702-0705台南團X14'
    sheet.append(['項目', '單價', '數量', '總價', '備註'])
    sheet.append(['機票', 3500, 14, 49000, '已付款'])
    sheet.append(['雙人房', 5000, 7, 35000, '已付訂金'])
    cancelled = workbook.create_sheet('20260303-0305測試團-取消')
    cancelled.append(['行程', '航班', '未付款'])
    reference = workbook.create_sheet('景點參考')
    reference.append(['景點', '價格'])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class GroupImportTest(unittest.TestCase):
    @classmethod
    def tearDownClass(cls):
        with app.app_context():
            db.session.remove()
            db.engine.dispose()
        _temp_dir.cleanup()

    def setUp(self):
        app.config.update(TESTING=True)
        with app.app_context():
            database_path = db.engine.url.database
            if database_path and database_path != ':memory:':
                os.makedirs(os.path.dirname(database_path), exist_ok=True)
            db.drop_all()
            db.create_all()
            admin = User(username='admin-import', role='admin', display_name='管理員')
            admin.set_password('password')
            viewer = User(username='viewer-import', role='viewer', display_name='檢視者')
            viewer.set_password('password')
            trip = Trip(
                customer_name='原始測試團', customer_phone='0912345678',
                trip_date='2026-07-02', return_date='2026-07-05',
                days=4, adults=14, total_people=14, status='確認',
                transport_cost=49000, accommodation_cost=35000,
                cost_subtotal=84000, final_quote=84000, quote_per_person=6000,
                itinerary_data='{}',
            )
            db.session.add_all([admin, viewer, trip])
            db.session.flush()
            operation = TripOperation(
                trip_id=trip.id, group_name='原始測試團', source_sheet='原始資料',
                payment_status='已付訂金', deposit_amount=10000,
                rooming_details='雙人房 7 間',
            )
            db.session.add(operation)
            db.session.commit()
            self.trip_id = trip.id
        self.client = app.test_client()

    def login(self, username='admin-import'):
        response = self.client.post('/api/auth/login', json={'username': username, 'password': 'password'})
        self.assertEqual(response.status_code, 200)

    def test_date_parser_handles_compact_cross_month_ranges(self):
        self.assertEqual(parse_sheet_dates('0529-0601測試', 2026), ('2026-05-29', '2026-06-01'))
        self.assertEqual(parse_sheet_dates('729-82測試', 2026), ('2026-07-29', '2026-08-02'))

    def test_preview_classifies_legacy_sheets_and_detects_signals(self):
        items = preview_workbook(legacy_workbook_bytes())
        self.assertEqual(len(items), 3)
        self.assertEqual(items[0]['kind'], 'trip')
        self.assertTrue(items[0]['selected'])
        self.assertEqual(items[0]['adults'], 14)
        self.assertIn('含付款資訊', items[0]['signals'])
        self.assertEqual(items[1]['kind'], 'cancelled')
        self.assertFalse(items[1]['selected'])
        self.assertEqual(items[2]['kind'], 'reference')

    def test_export_template_has_exact_import_headers(self):
        self.login()
        response = self.client.get('/api/trips/export/template')
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data), data_only=True)
        self.assertIn(STANDARD_SHEET, workbook.sheetnames)
        headers = [cell.value for cell in workbook[STANDARD_SHEET][1]]
        self.assertEqual(headers[:6], ['系統編號', '團名', '聯絡電話', '聯絡Email', '出發日期', '回程日期'])
        self.assertIn('欄位說明', workbook.sheetnames)

    def test_export_edit_and_reimport_updates_existing_trip(self):
        self.login()
        exported = self.client.get('/api/trips/export')
        self.assertEqual(exported.status_code, 200)
        workbook = load_workbook(io.BytesIO(exported.data))
        sheet = workbook[STANDARD_SHEET]
        self.assertEqual(sheet['A2'].value, self.trip_id)
        self.assertEqual(sheet['B2'].value, '原始測試團')
        self.assertEqual(sheet['X2'].value, '已付訂金')
        sheet['B2'] = '更新後測試團'
        sheet['K2'] = 50000
        buffer = io.BytesIO()
        workbook.save(buffer)

        preview = self.client.post('/api/trips/import/preview', data={
            'file': (io.BytesIO(buffer.getvalue()), 'standard.xlsx'),
        }, content_type='multipart/form-data')
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.json['items'][0]['trip_id'], self.trip_id)
        self.assertEqual(preview.json['items'][0]['signals'], ['標準格式'])

        imported = self.client.post('/api/trips/import', json={'items': preview.json['items']})
        self.assertEqual(imported.status_code, 201)
        self.assertEqual(imported.json['created_count'], 0)
        self.assertEqual(imported.json['updated_count'], 1)
        with app.app_context():
            trip = db.session.get(Trip, self.trip_id)
            self.assertEqual(trip.customer_name, '更新後測試團')
            self.assertEqual(trip.transport_cost, 50000)
            self.assertEqual(TripOperation.query.filter_by(trip_id=self.trip_id).count(), 1)

    def test_legacy_import_deduplicates_source_and_operations_update(self):
        self.login()
        preview = self.client.post('/api/trips/import/preview', data={
            'file': (io.BytesIO(legacy_workbook_bytes()), 'groups.xlsx'),
        }, content_type='multipart/form-data')
        imported = self.client.post('/api/trips/import', json={'items': preview.json['items']})
        self.assertEqual(imported.status_code, 201)
        self.assertEqual(imported.json['created_count'], 1)
        trip_id = imported.json['created'][0]['id']
        repeated = self.client.post('/api/trips/import', json={'items': preview.json['items']})
        self.assertEqual(repeated.json['created_count'], 0)
        self.assertEqual(len(repeated.json['skipped']), 1)

        updated = self.client.put(f'/api/trips/{trip_id}/operations', json={
            'group_name': '台南測試團', 'payment_status': '已付訂金',
            'deposit_amount': 10000, 'rooming_details': '雙人房 7 間',
        })
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.json['deposit_amount'], 10000)
        with app.app_context():
            self.assertEqual(db.session.get(Trip, trip_id).customer_name, '台南測試團')
            self.assertGreater(ActivityLog.query.filter_by(target_type='trip').count(), 0)

    def test_viewer_can_export_but_cannot_preview_or_update(self):
        self.login('viewer-import')
        self.assertEqual(self.client.get('/api/trips/export/template').status_code, 200)
        preview = self.client.post('/api/trips/import/preview', data={
            'file': (io.BytesIO(legacy_workbook_bytes()), 'groups.xlsx'),
        }, content_type='multipart/form-data')
        self.assertEqual(preview.status_code, 403)
        update = self.client.put(f'/api/trips/{self.trip_id}/operations', json={'group_name': '不可修改'})
        self.assertEqual(update.status_code, 403)


if __name__ == '__main__':
    unittest.main()
